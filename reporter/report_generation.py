import os
import asyncio
from datetime import datetime, timedelta
from typing import Dict, Any, List
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

class ReportGenerationAgent:
    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        self.workbook = None
        self.worksheet = None
        self.output_path = None
        self.current_row = 2  # 從第2行開始寫入資料（第1行是標題）
        self._write_lock = asyncio.Lock()  # 添加異步鎖以確保並行安全

        os.makedirs(output_dir, exist_ok=True)

    def initialize_excel_report(self) -> str:
        """
        初始化 Excel 報告檔案，建立標題行
        支援斷點續爬：如果當月報告已存在，載入現有檔案
        """
        # 使用年月格式作為檔案名稱，確保每月一個報告檔案
        current_month = datetime.now().strftime("%Y-%m")
        base_filename = f"website_summary_report_{current_month}"
        self.output_path = os.path.join(self.output_dir, f"{base_filename}.xlsx")
        
        # 檢查當月報告是否已存在
        if os.path.exists(self.output_path):
            print(f"📁 發現現有報告檔案: {self.output_path}")
            try:
                # 載入現有工作簿
                self.workbook = load_workbook(self.output_path)
                self.worksheet = self.workbook.active
                
                # 計算下一個要寫入的行號（最後一行的下一行）
                self.current_row = self.worksheet.max_row + 1
                
                print(f"✅ 已載入現有報告，將從第 {self.current_row} 行繼續寫入")
                return self.output_path
                
            except Exception as e:
                print(f"⚠️ 載入現有報告失敗: {e}")
                print("📝 將建立新的報告檔案")
                # 如果載入失敗，建立新檔案
                pass
        
        # 建立新的工作簿和工作表
        print(f"📝 建立新的報告檔案: {self.output_path}")
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = '網站統計摘要'
        
        # 設定標題行
        headers = [
            '網站名稱', '網站URL', '總頁面數', '有日期頁面數', '無日期頁面數',
            '最後更新日期', '一年前內容數量', '一年前內容比例(%)', 
            '失效內部頁面數', '失效外部連結數', '總外部連結數', '爬取耗時'
        ]
        
        for col, header in enumerate(headers, 1):
            self.worksheet.cell(row=1, column=col, value=header)
        
        # 設定欄位寬度
        column_widths = [25, 50, 12, 15, 15, 15, 15, 18, 15, 15, 15, 12]
        for col, width in enumerate(column_widths, 1):
            self.worksheet.column_dimensions[get_column_letter(col)].width = width
        
        # 儲存初始檔案
        self.workbook.save(self.output_path)
        
        return self.output_path
    
    def get_processed_urls(self) -> List[str]:
        """
        取得已處理的網站URL列表，用於斷點續爬
        """
        processed_urls = []
        
        if not self.workbook or not self.worksheet:
            print("⚠️ Excel 報告尚未初始化，無法取得已處理的URL列表")
            return processed_urls
        
        # 從第2行開始讀取（第1行是標題）
        for row in range(2, self.worksheet.max_row + 1):
            url_cell = self.worksheet.cell(row=row, column=2)  # 網站URL在第2列
            if url_cell.value:
                processed_urls.append(str(url_cell.value).strip())
        
        print(f"📋 發現 {len(processed_urls)} 個已處理的網站")
        return processed_urls
    
    async def add_site_to_excel(self, site_stats: Dict[str, Any], log_writer=None) -> None:
        """
        將單一網站的統計資料立即寫入 Excel 檔案
        """
        async with self._write_lock:  # 使用異步鎖確保並行安全
            if not self.workbook or not self.worksheet:
                raise ValueError("Excel 報告尚未初始化，請先呼叫 initialize_excel_report()")

            def _log(message: str):
                """統一的日誌輸出方法"""
                if log_writer:
                    log_writer.print_and_log(message)
                else:
                    print(message)

            site_name = site_stats['site_name']
            site_url = site_stats['site_url']
            crawl_results = site_stats['crawl_results']
            page_summary = site_stats['page_summary']
            external_link_results = site_stats['external_link_results']
            crawl_duration_formatted = site_stats.get('crawl_duration', '0分0秒')

            one_year_ago = datetime.now() - timedelta(days=365)

            # 統計總頁面數
            total_pages = len(crawl_results)

            # 統計失效的內部頁面數量（爬蟲失敗的頁面）
            failed_pages = 0

            for status in crawl_results:
                if status >= 400 or status == 0:
                    failed_pages += 1

            # 統計失效的外部連結數量
            failed_external_links = 0
            for link, link_info in external_link_results.items():
                if link_info.get('status', 0) >= 400 or link_info.get('status', 0) == 0:
                    failed_external_links += 1

            # 找到最新的更新日期
            latest_update = None
            outdated_pages = 0
            no_date_pages = 0
            past_dates = []  # 今天或以前的日期
            future_dates = []  # 未來日期
            today = datetime.now().date()

            for url, info in page_summary.items():
                last_updated = info.get('last_updated', '')

                # 統計無日期的頁面
                if last_updated == "[無日期]" or last_updated == "[爬取失敗]" or not last_updated:
                    no_date_pages += 1
                    continue

                try:
                    update_date = datetime.strptime(last_updated, "%Y-%m-%d")
                    update_date_only = update_date.date()

                    if update_date_only <= today:
                        past_dates.append(update_date)
                        # 檢查是否為一年前的內容
                        if update_date < one_year_ago:
                            outdated_pages += 1
                    else:
                        future_dates.append(update_date)

                except ValueError:
                    no_date_pages += 1
                    continue

            # 計算最新更新日期：優先使用過去日期的最新值，沒有才用最接近今天的未來日期
            if past_dates:
                latest_update = max(past_dates).strftime("%Y-%m-%d")
            elif future_dates:
                latest_update = min(future_dates).strftime("%Y-%m-%d")
            else:
                latest_update = "無有效日期"

            # 計算一年前內容的比例
            pages_with_date = len(past_dates) + len(future_dates)
            if pages_with_date > 0:
                outdated_percentage = (outdated_pages / pages_with_date) * 100
            else:
                outdated_percentage = 0

            # 準備要寫入的資料
            row_data = [
                site_name,
                site_url,
                total_pages,
                pages_with_date,
                no_date_pages,
                latest_update,
                outdated_pages,
                round(outdated_percentage, 2),
                failed_pages,
                failed_external_links,
                len(external_link_results),
                crawl_duration_formatted
            ]

            # 寫入到當前行
            for col, value in enumerate(row_data, 1):
                self.worksheet.cell(row=self.current_row, column=col, value=value)

            # 移動到下一行
            self.current_row += 1

            # 立即儲存檔案
            self.workbook.save(self.output_path)

            _log(f"已將 '{site_name}' 的資料寫入 Excel (第 {self.current_row - 1} 行)")
    
    def finalize_excel_report(self) -> None:
        """
        完成 Excel 報告，進行最終儲存
        """
        if self.workbook:
            # 最終儲存
            self.workbook.save(self.output_path)
            # 關閉工作簿
            self.workbook.close()
            self.workbook = None
            self.worksheet = None